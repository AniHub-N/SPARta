#!/usr/bin/env python3
"""
read_excel.py — read a workbook, print what's in it, and flag the traps.

    python read_excel.py xlsx/Receivables_Ageing.xlsx              inventory + first rows
    python read_excel.py xlsx/Trial_Balance_by_Year.xlsx --sheet "TB 2019-20"
    python read_excel.py xlsx/Receivables_Ageing.xlsx --rows 50
    python read_excel.py xlsx/Receivables_Ageing.xlsx --csv out/    one CSV per sheet
    python read_excel.py xlsx/*.xlsx --inventory                    just the summary

WHY openpyxl AND NOT pandas OR calamine

    All three read these files correctly. openpyxl is the only one that tells you a cell
    holds a FORMULA rather than a value, and in this corpus that distinction is the whole
    ballgame:

        openpyxl              '=SUM(D2:D519)'
        openpyxl data_only    None
        pandas                NaN
        calamine              ''

    These workbooks were written by a script, not saved by Excel, so no formula result was
    ever cached in the file. Nothing has computed those sums. So:

      * data_only=True gives you None for every formula cell — 34 of them across the 9
        workbooks. Read a TOTAL row that way and None becomes 0 in your arithmetic.
      * pandas and calamine give you a blank, which is indistinguishable from an empty
        cell. You cannot tell a total row from a data row, so you may sum a column that
        already contains its own total, or drop rows you needed.

    With openpyxl the '=' prefix makes formula cells identifiable, so you can drop them and
    compute the totals yourself. That is what --check does below.

    Speed is not a consideration: all 9 workbooks load in 0.14s with openpyxl. calamine is
    ~100x faster and it does not matter at this size.
"""
import argparse, csv, glob, os, sys

is_formula = lambda v: isinstance(v, str) and v.startswith("=")


def inventory(path):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    out = []
    for ws in wb.worksheets:
        cells = [(c.coordinate, c.value) for row in ws.iter_rows() for c in row if is_formula(c.value)]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c) for c in rows[0]] if rows else []
        out.append(dict(sheet=ws.title, rows=ws.max_row, cols=ws.max_column,
                        formulas=cells, header=header, data=rows))
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("files", nargs="+", help="one or more .xlsx paths")
    ap.add_argument("--sheet", help="only this sheet")
    ap.add_argument("--rows", type=int, default=8, help="data rows to print (default: %(default)s)")
    ap.add_argument("--csv", metavar="DIR", help="write each sheet to DIR as CSV")
    ap.add_argument("--inventory", action="store_true", help="summary only, no rows")
    a = ap.parse_args()

    paths = [p for pat in a.files for p in sorted(glob.glob(pat))]
    if not paths:
        sys.exit("no such file(s)")

    for path in paths:
        print(f"\n{'=' * 74}\n{os.path.basename(path)}\n{'=' * 74}")
        for sh in inventory(path):
            if a.sheet and sh["sheet"] != a.sheet:
                continue
            print(f"\n-- sheet {sh['sheet']!r}  {sh['rows']} rows x {sh['cols']} cols  "
                  f"{len(sh['formulas'])} formula cell(s)")
            if sh["formulas"]:
                for coord, val in sh["formulas"]:
                    print(f"     FORMULA {coord}: {val}   <- no cached result; compute it yourself")
            if a.inventory:
                continue
            rows = sh["data"]
            if sh["header"]:
                print("   ", " | ".join(sh["header"]))
            for r in rows[1:1 + a.rows]:
                print("    ", ["" if v is None else v for v in r])
            if len(rows) - 1 > a.rows:
                print(f"     ... {len(rows) - 1 - a.rows} more rows")

            # the trap, made explicit
            body = [r for r in rows[1:] if not any(is_formula(v) for v in r)]
            dropped = len(rows) - 1 - len(body)
            if dropped:
                print(f"     NOTE {dropped} row(s) contain formulas (a totals row). Excluded from"
                      f" the {len(body)} data rows below.")
            for i, name in enumerate(sh["header"]):
                vals = [r[i] for r in body if isinstance(r[i], (int, float))]
                if len(vals) > 2:
                    print(f"     sum of {name!r} over {len(vals)} data rows = {sum(vals):,}")

            if a.csv:
                os.makedirs(a.csv, exist_ok=True)
                out = os.path.join(a.csv, f"{os.path.basename(path)[:-5]}__{sh['sheet']}.csv")
                with open(out, "w", newline="") as f:
                    csv.writer(f).writerows([["" if v is None else v for v in r] for r in rows])
                print(f"     wrote {out}")


if __name__ == "__main__":
    sys.exit(main())
