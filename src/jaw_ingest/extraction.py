from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import fitz
from openpyxl import load_workbook

from .cache import CacheManager
from .models import (
    Cell,
    Coordinate,
    DocumentType,
    EvidenceBlock,
    Page,
    PDFDocument,
    Sheet,
    Workbook,
)

logger = logging.getLogger(__name__)


def _normalize_document_type(path: Path) -> DocumentType:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return DocumentType.pdf
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return DocumentType.xlsx
    return DocumentType.unknown


def _build_document_metadata(path: Path, doc_type: DocumentType) -> dict[str, Any]:
    return {
        "filename": path.name,
        "document_type": doc_type.value,
        "size_bytes": path.stat().st_size,
    }


def _serialize_document(document: PDFDocument) -> dict[str, Any]:
    return document.model_dump(mode="json")


def _deserialize_document(payload: dict[str, Any]) -> PDFDocument:
    return PDFDocument.model_validate(payload)


def _serialize_workbook(workbook: Workbook) -> dict[str, Any]:
    return workbook.model_dump(mode="json")


def _deserialize_workbook(payload: dict[str, Any]) -> Workbook:
    return Workbook.model_validate(payload)


def extract_pdf_document(path: Path, cache_manager: CacheManager | None = None) -> PDFDocument:
    document_id = path.stem
    document_type = _normalize_document_type(path)
    metadata = _build_document_metadata(path, document_type)
    cache_key = f"pdf:{path.resolve()}"

    if cache_manager:
        cached = cache_manager.get(cache_key)
        if cached is not None:
            logger.debug("Using cached PDF extraction for %s", path)
            return _deserialize_document(cached)

    document = PDFDocument(
        document_id=document_id,
        filename=path.name,
        document_type=document_type,
        source_path=path.resolve(),
        metadata=metadata,
    )

    with fitz.open(path) as doc:
        for page_number in range(len(doc)):
            page = doc[page_number]
            blocks: list[EvidenceBlock] = []
            image_paths: list[Path] = []
            text = page.get_text("text")
            if not text.strip():
                logger.debug("Page %s of %s contains no text; rasterizing page image.", page_number + 1, path)
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                image_path = path.with_name(f"{document_id}_page{page_number + 1}.png")
                pix.save(image_path)
                image_paths.append(image_path)

            for block_index, block in enumerate(page.get_text("blocks"), start=1):
                if len(block) < 5:
                    logger.warning(
                        "Unexpected block format on %s page %s: %s",
                        path,
                        page_number + 1,
                        block,
                    )
                    continue 

                x0, y0, x1, y1, block_text = block[:5]
                blocks.append(
                    EvidenceBlock(
                        block_id=f"{document_id}-p{page_number + 1}-b{block_index}",
                        document_id=document_id,
                        page_number=page_number + 1,
                        text=str(block_text).strip(),
                        coordinates=Coordinate(x0=x0, y0=y0, x1=x1, y1=y1),
                    )
                )

            document.pages.append(
                Page(
                    page_number=page_number + 1,
                    text=text,
                    blocks=blocks,
                    embedded_image_paths=image_paths,
                )
            )

    if cache_manager:
        cache_manager.set(cache_key, _serialize_document(document))
    return document


def extract_xlsx_workbook(path: Path, cache_manager: CacheManager | None = None) -> Workbook:
    workbook_id = path.stem
    document_type = _normalize_document_type(path)
    metadata = _build_document_metadata(path, document_type)
    cache_key = f"xlsx:{path.resolve()}"

    if cache_manager:
        cached = cache_manager.get(cache_key)
        if cached is not None:
            logger.debug("Using cached workbook extraction for %s", path)
            return _deserialize_workbook(cached)

    workbook = Workbook(
        document_id=workbook_id,
        filename=path.name,
        document_type=document_type,
        source_path=path.resolve(),
        metadata=metadata,
    )

    workbook_obj = load_workbook(path, data_only=False, read_only=False)
    for sheet_name in workbook_obj.sheetnames:
        sheet_obj = workbook_obj[sheet_name]
        sheet = Sheet(workbook_id=workbook_id, sheet_name=sheet_name)

        for row in sheet_obj.iter_rows(values_only=False):
            for cell_obj in row:
                formula = cell_obj.value if cell_obj.data_type == "f" else None
                comment = getattr(cell_obj, "comment", None)
                cell = Cell(
                    workbook_id=workbook_id,
                    sheet_name=sheet_name,
                    cell_address=cell_obj.coordinate,
                    value=cell_obj.value,
                    formula=formula,
                    note=comment.text if comment is not None else None,
                    metadata={
                        "data_type": cell_obj.data_type,
                        "number_format": cell_obj.number_format,
                    },
                )
                sheet.cells.append(cell)

        workbook.sheets.append(sheet)

    workbook_obj.close()

    if cache_manager:
        cache_manager.set(cache_key, _serialize_workbook(workbook))
    return workbook


def extract_document(path: Path, cache_manager: CacheManager | None = None) -> PDFDocument | Workbook:
    if path.suffix.lower() == ".pdf":
        return extract_pdf_document(path, cache_manager=cache_manager)
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return extract_xlsx_workbook(path, cache_manager=cache_manager)
    raise ValueError(f"Unsupported document type: {path.suffix}")
